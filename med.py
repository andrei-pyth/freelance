import openpyxl


class Cases():
    def __init__(self, sheet, number):
        self.id = number
        self.mkb = sheet.cell(row=number, column=8).value
        self.age = sheet.cell(row=number, column=10).value
        self.opernumber = sheet.cell(row=number, column=11).value

def main():
    doc = openpyxl.load_workbook('med.xlsx')
    sheet = doc.get_sheet_by_name('Лист1')
    sheet_active = doc.active
    rows = sheet.max_row
    compl_lst = []
    cases_lst = []
    for item in range(2, rows+1):
        res7 = Cases(sheet, item)
        sheet_active.cell(row=item, column=18).value = analyt(res7)
    doc.save('med.xlsx')

def analyt(obj):
    if obj.opernumber <= 924.5:
        if obj.opernumber <= 78.65:
            return 'нет'
        else:
            if obj.age <= 75.1:
                if obj.opernumber <= 159:
                    return 'нет'
                else:
                    if obj.opernumber <= 322:
                        return 'есть'
                    else:
                        if obj.opernumber <= 497:
                            return 'нет'
                        else:
                            return 'есть'
            else: return 'есть'
    else:
        if obj.age <= 75:
            if obj.age <= 73:
                if obj.opernumber <= 1444:
                    if obj.mkb == 'Старческая морганиева катаракта':
                        return 'есть'
                    else:
                        if obj.age <= 59:
                            return 'нет'
                        else:
                            return 'есть'
                else:
                    if obj.age <= 37:
                        return 'есть'
                    else:
                        return 'нет'
            else:
                if obj.mkb == 'Старческая ядерная катаракта':
                    return 'нет'
                else:
                    if obj.age <= 80:
                        if obj.mkb == 'Начальная старческая катаракта':
                            return 'нет'
                        else:
                            if obj.age <= 76:
                                return 'есть'
                            else:
                                if obj.age <= 77:
                                    return 'нет'
                                else:
                                    return 'есть'
        else:
            if obj.mkb == 'старческая ядерная катаракта':
                return 'нет'
            else: return 'есть'
